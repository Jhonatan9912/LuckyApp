# app/routes/admin/admin_routes.py
import io
from datetime import datetime, timezone
from flask import jsonify, current_app, Response, send_file
from flask_jwt_extended import jwt_required, get_jwt, get_jwt_identity
from sqlalchemy import text
from app.db.database import db
from . import bp
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from app.services.admin.admin_service import (
    get_lottery_dashboard_summary,
    get_active_games_export_rows,
)


def _require_admin():
    """
    Valida que el usuario actual tenga role_id = 1 (admin).
    Lanza una respuesta 403 si no es admin.
    """
    try:
        claims = get_jwt() or {}
        role_id = claims.get("role_id")

        if role_id is None:
            uid = int(get_jwt_identity())
            role_id = db.session.execute(
                text("SELECT role_id FROM users WHERE id = :uid"),
                {"uid": uid},
            ).scalar()

        if int(role_id) != 1:
            return jsonify({"ok": False, "error": "Solo administradores"}), 403

        return None  # OK
    except Exception:
        current_app.logger.exception("_require_admin: role check failed")
        return jsonify({"ok": False, "error": "Rol inválido"}), 403


# -------------------------------------------------------------------
#  RESUMEN DEL DASHBOARD
# -------------------------------------------------------------------
@bp.get("/dashboard/summary")
@jwt_required()
def dashboard_summary():
    # 1) Validar admin
    resp = _require_admin()
    if resp is not None:
        return resp

    # 2) Devolver el mismo shape que usaba Flutter
    try:
        summary = get_lottery_dashboard_summary()

        # Por si el service devuelve (data, status)
        if isinstance(summary, tuple):
            summary = summary[0]

        if not isinstance(summary, dict):
            summary = {"value": summary}

        return jsonify(summary), 200
    except Exception:
        current_app.logger.exception("dashboard_summary: service error")
        return (
            jsonify(
                {"ok": False, "error": "Fallo en dashboard_summary"},
            ),
            500,
        )


# -------------------------------------------------------------------
#  EXPORTAR JUEGOS ACTIVOS + NÚMEROS RESERVADOS (Excel .xlsx)
# -------------------------------------------------------------------
@bp.get("/dashboard/export-active-games")
@jwt_required()
def export_active_games():
    resp = _require_admin()
    if resp is not None:
        return resp

    try:
        rows = get_active_games_export_rows()
    except Exception:
        current_app.logger.exception("export_active_games: service error")
        return jsonify({"ok": False, "error": "No se pudo obtener la información"}), 500

    # ── Estilos ─────────────────────────────────────────────────────
    GOLD      = "FFD700"
    DARK      = "1A1A2E"
    BLUE_HDR  = "16213E"
    BLUE_ROW  = "0F3460"
    WHITE     = "FFFFFF"
    GRAY_ALT  = "F2F2F2"
    GRAY_DARK = "D9D9D9"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=WHITE, size=11):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Libro ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Juegos Activos"

    # Fila 1: título general del reporte
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de Juegos Activos  —  {generated}"
    title_cell.font      = Font(bold=True, color=GOLD, size=14, name="Calibri")
    title_cell.fill      = _fill(DARK)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 32

    ws.append([])  # fila 2 vacía como separador

    current_game_id = None
    row_idx = 3   # empezamos en fila 3

    if not rows:
        ws.merge_cells(f"A{row_idx}:E{row_idx}")
        ws[f"A{row_idx}"].value     = "No hay juegos activos ni números reservados."
        ws[f"A{row_idx}"].font      = _font(color="666666")
        ws[f"A{row_idx}"].alignment = center
    else:
        alt = False  # alternancia de color en filas de datos

        for r in rows:
            game_id   = r.get("game_id")
            lot_name  = r.get("lottery_name") or "—"
            p_date    = r.get("played_date")  or "—"
            p_time    = r.get("played_time")  or "—"
            players   = r.get("players_in_game") or 0
            reserved  = r.get("reserved_numbers_in_game") or 0
            digits    = int(r.get("digits") or 3)

            # ── Cabecera del juego ───────────────────────────────
            if game_id != current_game_id:
                if current_game_id is not None:
                    row_idx += 1  # línea en blanco entre juegos

                # Título del juego (fila fusionada)
                ws.merge_cells(f"A{row_idx}:E{row_idx}")
                c = ws[f"A{row_idx}"]
                c.value     = f"Juego #{game_id}  ·  {lot_name}  ·  {p_date}  {p_time}"
                c.font      = Font(bold=True, color=WHITE, size=12, name="Calibri")
                c.fill      = _fill(BLUE_HDR)
                c.alignment = left
                ws.row_dimensions[row_idx].height = 24
                row_idx += 1

                # Estadísticas (2 celdas)
                for col, label, val in [
                    (1, "Jugadores", players),
                    (3, "Números reservados", reserved),
                ]:
                    lc = ws.cell(row=row_idx, column=col)
                    lc.value     = label
                    lc.font      = Font(bold=True, color=GOLD, size=10, name="Calibri")
                    lc.fill      = _fill(BLUE_ROW)
                    lc.alignment = center
                    lc.border    = border

                    vc = ws.cell(row=row_idx, column=col + 1)
                    vc.value     = val
                    vc.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
                    vc.fill      = _fill(BLUE_ROW)
                    vc.alignment = center
                    vc.border    = border

                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

                # Encabezados de columnas
                headers = ["ID Usuario", "Nombre", "Teléfono", f"Número ({digits} dígitos)", "Suscripción"]
                for col, h in enumerate(headers, 1):
                    hc = ws.cell(row=row_idx, column=col)
                    hc.value     = h
                    hc.font      = Font(bold=True, color=DARK, size=10, name="Calibri")
                    hc.fill      = _fill(GOLD)
                    hc.alignment = center
                    hc.border    = border
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

                current_game_id = game_id
                alt = False

            # ── Fila de datos ────────────────────────────────────
            raw_num = r.get("number")
            if raw_num is not None:
                num_str = str(int(raw_num)).zfill(digits)
            else:
                num_str = "—"

            row_fill = _fill(GRAY_ALT) if alt else _fill(WHITE)
            font_data = Font(color="222222", size=10, name="Calibri")

            data = [
                r.get("user_id", ""),
                r.get("user_name", "") or "—",
                r.get("user_phone", "") or "—",
                num_str,
                "",  # columna suscripción (vacía, para uso futuro)
            ]
            for col, val in enumerate(data, 1):
                dc = ws.cell(row=row_idx, column=col)
                dc.value     = val
                dc.font      = font_data
                dc.fill      = row_fill
                dc.alignment = left
                dc.border    = border
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            alt = not alt

    # ── Anchos de columna ────────────────────────────────────────────
    col_widths = [12, 30, 18, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congelar la fila de título
    ws.freeze_panes = "A3"

    # ── Serializar y devolver ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"juegos_activos_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
